import cv2
import numpy as np
import base64
import os
import pandas as pd
from datetime import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_from_directory
from functools import wraps
from models import db, Student, Attendance, Faculty, Syllabus
from sqlalchemy import func, and_
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import Column, Integer, String, ForeignKey
from sqlalchemy.orm import relationship

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this to a secure secret key
# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///face_attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Constants from Kivy app
DEPARTMENTS = {
    "Microbiology": ["Dr. Chetan Chandrakant Ambasana", "Dr. Jignasha Trikamlal Thumar", 
                     "Dr. Arefa Abdulkhaliq Baakza", "Dr. Sheetal Pramodbhai Pithva", 
                     "Dr. Dharmesh N Adhyaru"],

    "Mathematics": ["Dr. Bhavin Mansukhla Patel", "Dr. Yogita Madhukant Parmar", 
                    "Mrs. Meena Mulchandani", "Mrs. Mamta S Amrutiya"],

    "Physics": ["Dr. Mukesh Vachhani", "Dr. Pooja Uttamprakash Sharma", 
                "Dr. Vibha B Vansola", "Dr. Kashmira P Tank", 
                "Mr. Prahlad Chaudhary", "Dr. Amitkumar Patel", "Mr. Pritesh Khatri"],

    "Botany": ["Dr. Suresh K Patel", "Dr. Nailesh A Patel", 
               "Dr. Pragna Parsottambhai Prajapati", "Dr. Mukesh M Patel", 
               "Dr. Rita Shivlal Chudasama", "Dr. Rohitkumar Patel", 
               "Dr. Binny Karlikar", "Mr. Chandresh R Kharadi"],

    "Zoology": ["Dr. Chetana V Shah", "Dr. Manishkumar D Visavadia", "Dr. Heena Prajapati"],

    "Chemistry": ["Dr. Shaileshkumar Prajapati", "Dr. K.B Patel", "Dr. K.S Nimavat", 
                  "Khachar Natubhai Bhurabhai", "Dr. Harshad P Lakum", 
                  "Dr. Chetan B Sangani", "Miss Nisha Rameshchand Sharma", "Dr. Mamta T Singh"],

    "Statistics": ["Dr. Samir Pandya", "Dr. Dharak Patel"]
}

SUBJECT_CODES = {
    "Microbiology": ["MB101", "MB102", "MB103"],
    "Mathematics": ["M101", "M102", "M103"],
    "Physics": ["P101", "P102", "P103"],
    "Botany": ["B101", "B102", "B103"],
    "Zoology": ["Z101", "Z102", "Z103"],
    "Chemistry": ["C101", "C102", "C103"],
    "Statistics": ["S101", "S102", "S103"]
}

UNITS = ["Unit 1", "Unit 2", "Unit 3", "Unit 4"]
SEMESTERS = ['Semester 1', 'Semester 2', 'Semester 3', 'Semester 4', 'Semester 5', 'Semester 6']

# Store registered faces and attendance
registered_faces = {}
attendance_records = []

# Faculty credentials
FACULTY_CREDENTIALS = {
    'Microbiology': {
        'CCA001': {'name': 'Dr. Chetan Chandrakant Ambasana', 'password': 'cca@123'},
        'JTT001': {'name': 'Dr. Jignasha Trikamlal Thumar', 'password': 'jtt@123'},
        'AAB001': {'name': 'Dr. Arefa Abdulkhaliq Baakza', 'password': 'aab@123'},
        'SPP001': {'name': 'Dr. Sheetal Pramodbhai Pithva', 'password': 'spp@123'},
        'DNA001': {'name': 'Dr. Dharmesh N Adhyaru', 'password': 'dna@123'}
    },
    'Mathematics': {
        'BMP001': {'name': 'Dr. Bhavin Mansukhla Patel', 'password': 'bmp@123'},
        'YMP001': {'name': 'Dr. Yogita Madhukant Parmar', 'password': 'ymp@123'},
        'MM001': {'name': 'Mrs. Meena Mulchandani', 'password': 'mm@123'},
        'MSA001': {'name': 'Mrs. Mamta S Amrutiya', 'password': 'msa@123'}
    },
    'Physics': {
        'MV001': {'name': 'Dr. Mukesh Vachhani', 'password': 'mv@123'},
        'PUS001': {'name': 'Dr. Pooja Uttamprakash Sharma', 'password': 'pus@123'},
        'VBV001': {'name': 'Dr. Vibha B Vansola', 'password': 'vbv@123'},
        'KPT001': {'name': 'Dr. Kashmira P Tank', 'password': 'kpt@123'},
        'PC001': {'name': 'Mr. Prahlad Chaudhary', 'password': 'pc@123'},
        'AP001': {'name': 'Dr. Amitkumar Patel', 'password': 'ap@123'},
        'PK001': {'name': 'Mr. Pritesh Khatri', 'password': 'pk@123'}
    },
    'Botany': {
        'SKP001': {'name': 'Dr. Suresh K Patel', 'password': 'skp@123'},
        'NAP001': {'name': 'Dr. Nailesh A Patel', 'password': 'nap@123'},
        'PPP001': {'name': 'Dr. Pragna Parsottambhai Prajapati', 'password': 'ppp@123'},
        'MMP001': {'name': 'Dr. Mukesh M Patel', 'password': 'mmp@123'},
        'RSC001': {'name': 'Dr. Rita Shivlal Chudasama', 'password': 'rsc@123'},
        'RP001': {'name': 'Dr. Rohitkumar Patel', 'password': 'rp@123'},
        'BK001': {'name': 'Dr. Binny Karlikar', 'password': 'bk@123'},
        'CRK001': {'name': 'Mr. Chandresh R Kharadi', 'password': 'crk@123'}
    },
    'Zoology': {
        'CVS001': {'name': 'Dr. Chetana V Shah', 'password': 'cvs@123'},
        'MDV001': {'name': 'Dr. Manishkumar D Visavadia', 'password': 'mdv@123'},
        'HP001': {'name': 'Dr. Heena Prajapati', 'password': 'hp@123'}
    },
    'Chemistry': {
        'SP001': {'name': 'Dr. Shaileshkumar Prajapati', 'password': 'sp@123'},
        'KBP001': {'name': 'Dr. K.B Patel', 'password': 'kbp@123'},
        'KSN001': {'name': 'Dr. K.S Nimavat', 'password': 'ksn@123'},
        'KNB001': {'name': 'Khachar Natubhai Bhurabhai', 'password': 'knb@123'},
        'HPL001': {'name': 'Dr. Harshad P Lakum', 'password': 'hpl@123'},
        'CBS001': {'name': 'Dr. Chetan B Sangani', 'password': 'cbs@123'},
        'NRS001': {'name': 'Miss Nisha Rameshchand Sharma', 'password': 'nrs@123'},
        'MTS001': {'name': 'Dr. Mamta T Singh', 'password': 'mts@123'}
    },
    'Statistics': {
        'SP002': {'name': 'Dr. Samir Pandya', 'password': 'sp@123'},
        'DP001': {'name': 'Dr. Dharak Patel', 'password': 'dp@123'}
    }
}

# Admin credentials (you should move these to a secure configuration file in production)
ADMIN_CREDENTIALS = {
    'username': 'admin',
    'password': 'admin123'  # In production, use a strong hashed password
}

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'faculty_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin session management
def admin_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login_page'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        faculty = Faculty.query.filter_by(username=username).first()
        
        if faculty and faculty.check_password(password):
            session['faculty_id'] = faculty.id
            session['faculty_name'] = faculty.name
            session['department'] = faculty.department
            return redirect(url_for('index'))
            
        return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html', 
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/get_faculty/<department>')
def get_faculty(department):
    faculty_list = DEPARTMENTS.get(department, [])
    return jsonify(faculty_list)

@app.route('/get_subjects/<department>')
def get_subjects(department):
    subject_list = SUBJECT_CODES.get(department, [])
    return jsonify(subject_list)

@app.route('/register', methods=['POST'])
@admin_login_required
def register():
    try:
        data = request.json
        if not data or 'images' not in data or len(data['images']) != 3:
            return jsonify({'success': False, 'message': 'Three photos are required for registration'})

        required_fields = ['name', 'enrollment', 'semester', 'group', 'major_subject', 'minor_subject', 'multi_subject']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Missing required field: {field}'})

        # Process all three images and store their encodings
        face_encodings = []
        photos = []
        
        for img_data in data['images']:
            img = process_base64_image(img_data)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            if len(faces) == 0:
                return jsonify({'success': False, 'message': 'No face detected in one or more photos'})
            
            if len(faces) > 1:
                    return jsonify({'success': False, 'message': 'Multiple faces detected in one photo. Please capture one face at a time'})
            
            # Convert image to binary for storage
            _, img_encoded = cv2.imencode('.jpg', img)
            photos.append(img_encoded.tobytes())
            face_encodings.append(faces[0].tolist())

        # Create new student record
        new_student = Student(
            name=data['name'],
            enrollment_number=data['enrollment'],
            semester=data['semester'],
            group=data['group'],
            major_subject=data['major_subject'],
            minor_subject=data['minor_subject'],
            multi_subject=data['multi_subject'],
            photo=photos[0],  # Store the first photo as main photo
            face_encoding=face_encodings  # Store all face encodings
        )

        # Add and commit to database
        db.session.add(new_student)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully registered {new_student.name} with 3 photos'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Registration error: {str(e)}")
        return jsonify({'success': False, 'message': f'Error during registration: {str(e)}'})

@app.route('/mark-attendance', methods=['POST'])
@login_required
def mark_attendance():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({'success': False, 'message': 'No image data provided'})

        # Process image
        img = process_base64_image(data['image'])

        # Detect faces
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        if len(faces) == 0:
            return jsonify({'success': False, 'message': 'No faces detected'})

        # Get students from the selected semester only
        students = Student.query.filter_by(semester=data['semester']).all()
        if not students:
            return jsonify({
                'success': False,
                'message': f'No students found in {data["semester"]}'
            })

        recognized_students = set()  # Use set to avoid duplicate attendance

        # For each detected face, check against registered students from the selected semester
        for (x, y, w, h) in faces:
            face_encoding = [x, y, w, h]  # In a real implementation, you'd use proper face recognition here
            
            for student in students:
                # In a real implementation, you'd compare face_encoding with student.face_encoding
                # For now, we'll just simulate recognition
                if student.id not in recognized_students:
                    attendance = Attendance(
                        student_id=student.id,
                        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
                        time=datetime.strptime(data['time'], '%H:%M').time(),
                        department=data['department'],
                        faculty=data['faculty'],
                        subject=data['subject'],
                        unit=data['unit']
                    )
                    db.session.add(attendance)
                    recognized_students.add(student.id)

        if recognized_students:
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Recorded attendance for {len(recognized_students)} students from {data["semester"]}',
                'recognized_students': [{'enrollment_number': student.enrollment_number, 'name': student.name} for student in students if student.id in recognized_students]
            })
        else:
            return jsonify({
                'success': False,
                'message': f'No registered students from {data["semester"]} were recognized'
            })

    except Exception as e:
        db.session.rollback()
        print(f"Attendance marking error: {str(e)}")
        return jsonify({'success': False, 'message': f'Error marking attendance: {str(e)}'})

@app.route('/export-attendance', methods=['POST'])
@login_required
def export_attendance():
    try:
        department = request.json.get('department', '')
        if not department:
            return jsonify({
                'success': False,
                'message': 'Department is required'
            })

        # Create exports directory if it doesn't exist
        current_dir = os.path.dirname(os.path.abspath(__file__))
        exports_dir = os.path.join(current_dir, 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Define fixed filename for each department in exports directory
        try:
            filename = os.path.join(exports_dir, f"attendance_{department}.xlsx")
            print(f"Will save Excel file to: {filename}")
        except Exception as e:
            print(f"Error creating file path: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to create file path: {str(e)}'
            })

        # Query attendance records
        try:
            attendance_data = (
                db.session.query(
                    Student.name,
                    Student.enrollment_number,
                    Student.semester,
                    Attendance.date,
                    Attendance.time,
                    Attendance.department,
                    Attendance.faculty,
                    Attendance.subject,
                    Attendance.unit
                )
                .select_from(Attendance)
                .join(Student, Student.id == Attendance.student_id)
                .filter(Attendance.department == department)
                .order_by(
                    Attendance.date.desc(),
                    Attendance.time.desc(),
                    Student.enrollment_number
                )
                .all()
            )
            print(f"Found {len(attendance_data)} attendance records")
        except Exception as e:
            print(f"Database query error: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to query attendance data: {str(e)}'
            })

        if not attendance_data:
            print("No attendance records found")
            return jsonify({
                'success': False,
                'message': 'No attendance records found for the selected department'
            })

        # Convert to DataFrame
        try:
            records = []
            for record in attendance_data:
                record_dict = {
                    'Enrollment Number': record.enrollment_number,
                    'Student Name': record.name,
                    'Semester': record.semester,
                    'Date': record.date.strftime('%Y-%m-%d'),
                    'Time': record.time.strftime('%H:%M'),
                    'Faculty': record.faculty,
                    'Subject': record.subject,
                    'Unit': record.unit
                }
                records.append(record_dict)
                print(f"Processing record: {record_dict}")

            df = pd.DataFrame(records)
            df = df.sort_values(['Date', 'Time', 'Enrollment Number'], 
                              ascending=[False, False, True])

            column_order = [
                'Enrollment Number',
                'Student Name',
                'Semester',
                'Date',
                'Time',
                'Faculty',
                'Subject',
                'Unit'
            ]
            df = df[column_order]
            print(f"Created DataFrame with shape: {df.shape}")
        except Exception as e:
            print(f"Error creating DataFrame: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to process attendance data: {str(e)}'
            })

        # Save Excel file
        try:
            # First try to remove the file if it exists
            if os.path.exists(filename):
                os.remove(filename)
                print(f"Removed existing file: {filename}")

            # Save the new file
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Attendance Records')
                workbook = writer.book
                worksheet = writer.sheets['Attendance Records']

                # Format header
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Format cells
                for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    fill_color = 'F0F8FF' if (idx - 2) % 2 == 0 else 'FFFFFF'
                    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        cell.fill = row_fill

                # Adjust column widths
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

            print(f"Successfully saved Excel file to: {filename}")
            return jsonify({
                'success': True,
                'message': f'Attendance exported to exports/attendance_{department}.xlsx'
            })

        except Exception as e:
            print(f"Error saving Excel file: {str(e)}")
            return jsonify({
                'success': False,
                'message': f'Failed to save Excel file: {str(e)}'
            })

    except Exception as e:
        print(f"General export error: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to export attendance: {str(e)}'
        })

def process_base64_image(base64_string):
    try:
        # Remove the data URL prefix if present
        if ',' in base64_string:
            base64_string = base64_string.split(',')[1]
        
        # Decode base64 string to bytes
        image_bytes = base64.b64decode(base64_string)
        
        # Convert bytes to numpy array
        nparr = np.frombuffer(image_bytes, np.uint8)
        
        # Decode numpy array as image
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            raise ValueError("Failed to decode image")
            
        return image
    except Exception as e:
        print(f"Error processing base64 image: {str(e)}")
        raise

@app.route('/register-student')
@admin_login_required
def register_student():
    return render_template('register.html',
                         departments=DEPARTMENTS,
                         subject_codes=SUBJECT_CODES,
                         units=UNITS,
                         semesters=SEMESTERS)

@app.route('/mark-attendance-page')
@login_required
def mark_attendance_page():
    return render_template('attendance.html',
                         departments=DEPARTMENTS,
                         semesters=SEMESTERS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/export-attendance-page')
@login_required
def export_attendance_page():
    return render_template('export.html',
                         departments=DEPARTMENTS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/get-semester-students/<semester>')
@login_required
def get_semester_students(semester):
    try:
        students = Student.query.filter_by(semester=semester).all()
        student_list = [{
            'id': student.id,
            'name': student.name,
            'enrollment_number': student.enrollment_number
        } for student in students]
        
        return jsonify({
            'success': True,
            'students': student_list
        })
    except Exception as e:
        print(f"Error fetching students: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching students: {str(e)}'
        })

@app.route('/mark-student-present', methods=['POST'])
@login_required
def mark_student_present():
    try:
        data = request.json
        if not data or 'student_id' not in data:
            return jsonify({
                'success': False,
                'message': 'Student ID is required'
            })

        student = Student.query.get(data['student_id'])
        if not student:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            })

        # Create attendance record
        attendance = Attendance(
            student_id=student.id,
            date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
            time=datetime.strptime(data['time'], '%H:%M').time(),
            department=data['department'],
            faculty=data['faculty'],
            subject=data['subject'],
            unit=data['unit']
        )
        db.session.add(attendance)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Marked {student.name} as present',
            'name': student.name,
            'enrollment_number': student.enrollment_number
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error marking student present: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error marking student present: {str(e)}'
        })

@app.route('/get-attendance-history', methods=['POST'])
@login_required
def get_attendance_history():
    try:
        data = request.json
        query = db.session.query(
            Attendance.date,
            Attendance.time,
            Student.semester,
            Attendance.subject,
            Attendance.unit,
            func.count(Attendance.student_id).label('present_count')
        ).join(
            Student,
            Student.id == Attendance.student_id
        ).group_by(
            Attendance.date,
            Attendance.time,
            Student.semester,
            Attendance.subject,
            Attendance.unit
        ).filter(
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        )

        # Apply filters if provided
        if data.get('startDate') and data.get('endDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date.between(start_date, end_date))
        elif data.get('startDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date >= start_date)
        elif data.get('endDate'):
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date <= end_date)

        if data.get('semester'):
            query = query.filter(Student.semester == data['semester'])
        if data.get('subject'):
            query = query.filter(Attendance.subject == data['subject'])

        records = query.order_by(Attendance.date.desc(), Attendance.time.desc()).all()

        # Calculate absent count for each record
        formatted_records = []
        for record in records:
            # Get total students in the semester
            total_students = Student.query.filter_by(semester=record.semester).count()
            
            record_dict = {
                'date': record.date.strftime('%Y-%m-%d'),
                'time': record.time.strftime('%H:%M'),
                'semester': record.semester,
                'subject': record.subject,
                'unit': record.unit,
                'present_count': record.present_count,
                'absent_count': total_students - record.present_count
            }
            formatted_records.append(record_dict)

        return jsonify({
            'success': True,
            'records': formatted_records
        })

    except Exception as e:
        print(f"Error fetching attendance history: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching attendance history: {str(e)}'
        })

@app.route('/attendance-details')
@login_required
def attendance_details():
    try:
        date = request.args.get('date')
        semester = request.args.get('semester')
        subject = request.args.get('subject')
        
        if not all([date, semester, subject]):
            print("Missing required parameters")
            return redirect(url_for('attendance_history_page'))

        # Get attendance records for the specific date, semester, and subject
        attendance_records = db.session.query(
            Student.name,
            Student.enrollment_number,
            Attendance.time,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.date == datetime.strptime(date, '%Y-%m-%d').date(),
            Student.semester == semester,
            Attendance.subject == subject,
            Attendance.faculty == session.get('faculty_name'),
            Attendance.department == session.get('department')
        ).all()

        # Get all students in the semester
        all_students = Student.query.filter_by(semester=semester).all()
        
        # Create a set of present student enrollment numbers
        present_enrollments = {record.enrollment_number for record in attendance_records}
        
        # Create present and absent lists
        present_students = attendance_records
        absent_students = [student for student in all_students 
                         if student.enrollment_number not in present_enrollments]

        # Get unique units from attendance records
        units = sorted(list(set(record.unit for record in attendance_records)))

        return render_template(
            'attendance_details.html',
            date=date,
            semester=semester,
            subject=subject,
            present_students=present_students,
            absent_students=absent_students,
            units=units,
            faculty_name=session.get('faculty_name'),
            department=session.get('department')
        )

    except Exception as e:
        print(f"Error in attendance_details: {str(e)}")
        return redirect(url_for('attendance_history_page'))

@app.route('/exports/<path:filename>')
@login_required
def serve_export(filename):
    try:
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        return send_from_directory(exports_dir, filename, as_attachment=True)
    except Exception as e:
        print(f"Error serving export: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error serving export: {str(e)}'
        })

@app.route('/export-specific-attendance', methods=['POST'])
@login_required
def export_specific_attendance():
    try:
        data = request.json
        date = datetime.strptime(data['date'], '%Y-%m-%d').date()

        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Query attendance records
        attendance_records = db.session.query(
            Student.enrollment_number,
            Student.name,
            Attendance.time,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.date == date,
            Student.semester == data['semester'],
            Attendance.subject == data['subject'],
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        ).all()

        # Get all students in the semester
        all_students = Student.query.filter_by(semester=data['semester']).all()
        
        # Create records list
        records = []
        present_students = {record.enrollment_number: record for record in attendance_records}
        
        for student in all_students:
            record = present_students.get(student.enrollment_number)
            records.append({
                'Enrollment Number': student.enrollment_number,
                'Student Name': student.name,
                'Status': 'Present' if record else 'Absent',
                'Time': record.time.strftime('%H:%M') if record else '-',
                'Unit': record.unit if record else '-'
            })

        # Create DataFrame and save to Excel
        df = pd.DataFrame(records)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"attendance_{data['department']}_{data['semester']}_{data['subject']}_{date}_{timestamp}.xlsx"
        filepath = os.path.join(exports_dir, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance')
            workbook = writer.book
            worksheet = writer.sheets['Attendance']

            # Format header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format status column
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, min_col=3, max_col=3), start=2):
                cell = row[0]
                if cell.value == 'Present':
                    cell.font = Font(color='008000')  # Green
                else:
                    cell.font = Font(color='FF0000')  # Red

            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

        return jsonify({
            'success': True,
            'message': 'Attendance exported successfully',
            'file_path': filename
        })

    except Exception as e:
        print(f"Error exporting specific attendance: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error exporting attendance: {str(e)}'
        })

@app.route('/export-all-attendance', methods=['POST'])
@login_required
def export_all_attendance():
    try:
        data = request.json
        
        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Build query with filters
        query = db.session.query(
            Student.name,
            Student.enrollment_number,
            Student.semester,
            Attendance.date,
            Attendance.time,
            Attendance.subject,
            Attendance.unit
        ).join(
            Attendance,
            Student.id == Attendance.student_id
        ).filter(
            Attendance.faculty == data['faculty'],
            Attendance.department == data['department']
        )

        # Apply filters if provided
        if data.get('startDate') and data.get('endDate'):
            start_date = datetime.strptime(data['startDate'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['endDate'], '%Y-%m-%d').date()
            query = query.filter(Attendance.date.between(start_date, end_date))
        if data.get('semester'):
            query = query.filter(Attendance.semester == data['semester'])
        if data.get('subject'):
            query = query.filter(Attendance.subject == data['subject'])

        # Execute query
        attendance_records = query.order_by(
            Attendance.date.desc(),
            Attendance.time.desc(),
            Student.enrollment_number
        ).all()

        if not attendance_records:
            return jsonify({
                'success': False,
                'message': 'No attendance records found for the selected criteria'
            })

        # Create DataFrame
        records = []
        for record in attendance_records:
            records.append({
                'Date': record.date.strftime('%Y-%m-%d'),
                'Time': record.time.strftime('%H:%M'),
                'Enrollment Number': record.enrollment_number,
                'Student Name': record.name,
                'Semester': record.semester,
                'Subject': record.subject,
                'Unit': record.unit
            })

        df = pd.DataFrame(records)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"attendance_report_{data['department']}_{timestamp}.xlsx"
        filepath = os.path.join(exports_dir, filename)

        # Save to Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance Records')
            workbook = writer.book
            worksheet = writer.sheets['Attendance Records']

            # Format header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format cells
            for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill_color = 'F0F8FF' if (idx - 2) % 2 == 0 else 'FFFFFF'
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = row_fill

            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

        return jsonify({
            'success': True,
            'message': 'Attendance report generated successfully',
            'file_path': filename
        })

    except Exception as e:
        print(f"Error exporting attendance report: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error generating report: {str(e)}'
        })

@app.route('/admin')
def admin_login_page():
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    try:
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            
            if not username or not password:
                return render_template('admin_login.html', error='Please provide both username and password')
            
            if username == ADMIN_CREDENTIALS['username'] and password == ADMIN_CREDENTIALS['password']:
                session['admin_logged_in'] = True
                return redirect(url_for('admin_dashboard'))
            
            return render_template('admin_login.html', error='Invalid credentials')
        
        # If it's a GET request, redirect to the login page
        return redirect(url_for('admin_login_page'))
    except Exception as e:
        print(f"Admin login error: {str(e)}")
        return render_template('admin_login.html', error='An error occurred. Please try again.')

@app.route('/admin/dashboard')
@admin_login_required
def admin_dashboard():
    try:
        # Get total counts
        total_faculty = Faculty.query.count()
        total_students = Student.query.count()
        
        # Get department-wise counts
        department_stats = {}
        for dept in DEPARTMENTS.keys():
            faculty_count = Faculty.query.filter_by(department=dept).count()
            student_count = Student.query.filter(
                Student.major_subject.like(f"%{dept}%")
            ).count()
            department_stats[dept] = {
                'faculty': faculty_count,
                'students': student_count
            }
        
        return render_template('admin_dashboard.html',
                             total_faculty=total_faculty,
                             total_students=total_students,
                             department_stats=department_stats)
    except Exception as e:
        print(f"Admin dashboard error: {str(e)}")
        return redirect(url_for('admin_login_page'))

@app.route('/admin/logout')
def admin_logout():
    try:
        session.pop('admin_logged_in', None)
        return redirect(url_for('admin_login_page'))
    except Exception as e:
        print(f"Admin logout error: {str(e)}")
        return redirect(url_for('admin_login_page'))

@app.route('/admin/manage-faculty')
@admin_login_required
def manage_faculty():
    try:
        faculty_members = Faculty.query.all()
        return render_template('manage_faculty.html', faculty_members=faculty_members)
    except Exception as e:
        print(f"Error fetching faculty members: {str(e)}")
        return render_template('manage_faculty.html', faculty_members=[], error="Error fetching faculty members")

@app.route('/admin/add-faculty', methods=['POST'])
@admin_login_required
def add_faculty():
    try:
        name = request.form.get('name')
        department = request.form.get('department')
        email = request.form.get('email')
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Check if username already exists
        if Faculty.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': 'Username already exists'})
        
        # Create new faculty member
        faculty = Faculty(
            name=name,
            department=department,
            email=email,
            username=username
        )
        faculty.set_password(password)
        
        db.session.add(faculty)
        db.session.commit()
        
        return redirect(url_for('manage_faculty'))
    except Exception as e:
        db.session.rollback()
        print(f"Error adding faculty: {str(e)}")
        return jsonify({'success': False, 'message': f'Error adding faculty: {str(e)}'})

@app.route('/admin/delete-faculty/<int:faculty_id>', methods=['POST'])
@admin_login_required
def delete_faculty(faculty_id):
    try:
        faculty = Faculty.query.get_or_404(faculty_id)
        db.session.delete(faculty)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting faculty: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting faculty: {str(e)}'})

@app.route('/admin/manage-syllabus')
@admin_login_required
def manage_syllabus():
    try:
        # Get all syllabus entries
        syllabus_entries = Syllabus.query.all()
        
        # Get unique departments from the constants
        departments = list(DEPARTMENTS.keys())
        
        return render_template('manage_syllabus.html',
                             departments=departments,
                             semesters=SEMESTERS,
                             syllabus_entries=syllabus_entries)
                             
    except Exception as e:
        print(f"Error in manage_syllabus route: {str(e)}")
        return render_template('manage_syllabus.html', 
                             departments=list(DEPARTMENTS.keys()),
                             semesters=SEMESTERS,
                             syllabus_entries=[],
                             error="Error fetching syllabus entries")

@app.route('/admin/add-syllabus', methods=['POST'])
@admin_login_required
def add_syllabus():
    try:
        department = request.form.get('department')
        semester = request.form.get('semester')
        subject = request.form.get('subject')
        units = request.form.getlist('units[]')  # Get multiple units
        
        # Join units with comma
        units_str = ','.join(units)
        
        # Check if entry already exists
        existing = Syllabus.query.filter_by(
            department=department,
            semester=semester,
            subject=subject
        ).first()
        
        if existing:
            # Update existing entry
            existing.units = units_str
            db.session.commit()
            return jsonify({'success': True, 'message': 'Syllabus updated successfully'})
        
        # Create new syllabus entry
        syllabus = Syllabus(
            department=department,
            semester=semester,
            subject=subject,
            units=units_str
        )
        
        db.session.add(syllabus)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Syllabus added successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error adding syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error adding syllabus: {str(e)}'})

@app.route('/admin/delete-syllabus/<int:syllabus_id>', methods=['POST'])
@admin_login_required
def delete_syllabus(syllabus_id):
    try:
        syllabus = Syllabus.query.get_or_404(syllabus_id)
        db.session.delete(syllabus)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Syllabus deleted successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error deleting syllabus: {str(e)}'})

@app.route('/admin/get-syllabus/<department>/<semester>')
@admin_login_required
def get_syllabus(department, semester):
    try:
        syllabus_entries = Syllabus.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        entries = [{
            'id': entry.id,
            'subject': entry.subject,
            'units': entry.units.split(',')
        } for entry in syllabus_entries]
        
        return jsonify({'success': True, 'entries': entries})
    except Exception as e:
        print(f"Error fetching syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error fetching syllabus: {str(e)}'})

@app.route('/admin/update-syllabus/<int:syllabus_id>', methods=['POST'])
@admin_login_required
def update_syllabus(syllabus_id):
    try:
        data = request.json
        syllabus = Syllabus.query.get_or_404(syllabus_id)
        
        # Update subject if provided
        if 'subject' in data:
            # Check if the new subject already exists for this department and semester
            existing = Syllabus.query.filter_by(
                department=syllabus.department,
                semester=syllabus.semester,
                subject=data['subject']
            ).first()
            
            if existing and existing.id != syllabus_id:
                return jsonify({
                    'success': False,
                    'message': 'This subject already exists for the selected department and semester'
                })
            
            syllabus.subject = data['subject']
        
        # Update units
        units = data.get('units', [])
        syllabus.units = ','.join(units)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Syllabus updated successfully'})
    except Exception as e:
        db.session.rollback()
        print(f"Error updating syllabus: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating syllabus: {str(e)}'})

# Create database tables and initialize faculty data
def init_faculty_data():
    try:
        # Check if faculty data already exists
        if Faculty.query.count() == 0:
            print("Initializing faculty data...")
            # Add faculty members from FACULTY_CREDENTIALS
            for department, faculty_dict in FACULTY_CREDENTIALS.items():
                for faculty_id, details in faculty_dict.items():
                    # Create faculty member
                    faculty = Faculty(
                        name=details['name'],
                        department=department,
                        email=f"{faculty_id.lower()}@college.edu",  # Generate email from faculty ID
                        username=faculty_id
                    )
                    faculty.set_password(details['password'])
                    db.session.add(faculty)
            
            db.session.commit()
            print("Faculty data initialized successfully!")
    except Exception as e:
        db.session.rollback()
        print(f"Error initializing faculty data: {str(e)}")

# Initialize database and faculty data
with app.app_context():
    db.create_all()
    init_faculty_data()

@app.route('/attendance-history')
@login_required
def attendance_history_page():
    return render_template('attendance_history.html',
                         departments=DEPARTMENTS,
                         subject_codes=SUBJECT_CODES,
                         units=UNITS,
                         semesters=SEMESTERS,
                         faculty_name=session.get('faculty_name'),
                         department=session.get('department'))

@app.route('/get-syllabus-for-attendance/<department>/<semester>')
@login_required
def get_syllabus_for_attendance(department, semester):
    try:
        syllabus_entries = Syllabus.query.filter_by(
            department=department,
            semester=semester
        ).all()
        
        entries = [{
            'id': entry.id,
            'subject': entry.subject,
            'units': entry.units.split(',')
        } for entry in syllabus_entries]
        
        return jsonify({
            'success': True,
            'entries': entries
        })
    except Exception as e:
        print(f"Error fetching syllabus for attendance: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching syllabus: {str(e)}'
        })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)